"""Parses a vendor-supplied "SO Details" Excel export into purchase-ready
rows, for the Purchases > Import from Vendor File feature.

Format this was built against: a two-sheet workbook ("SO Details" and
"Delivery Details") exported from a supplier's own order-management
system, with an "SO Details" sheet carrying columns SO Number, SO Date,
Material, Material Code, Quantity, Sales Unit, Net Value in Rs., Status.
Only "SO Details" is used (see app README) - it's the sheet that carries
a value per line, needed to work out a unit cost; "Delivery Details" (
batch numbers, no pricing) isn't used.

Like gst_logic.py's GSTR-2B parser, this is header-name-driven (scans for
the expected column headers by text rather than fixed positions) so a
harmless reordering of columns doesn't break it, and it fails loudly with
an explanatory message on a sheet/format it doesn't recognize rather than
guessing. This has only been built against the one sample file available
when it was written - a different vendor's export, or a future version of
this vendor's export, may need the header aliases below adjusted.
"""
import re

REQUIRED_HEADERS = {
    "so_number": ["so number"],
    "so_date": ["so date"],
    "material": ["material"],
    "material_code": ["material code"],
    "quantity": ["quantity"],
    "sales_unit": ["sales unit"],
    "net_value": ["net value"],
    "status": ["status"],
}


def _find_header_row(ws, max_scan_rows=5):
    max_row = min(max_scan_rows, ws.max_row or 0)
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        values = {}
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                values[c] = v.strip().lower()
        if not values:
            continue
        found = {}
        for key, aliases in REQUIRED_HEADERS.items():
            for c, v in values.items():
                if any(a in v for a in aliases):
                    found[key] = c
                    break
        # need at least these to consider it a match
        if all(k in found for k in ("material_code", "quantity", "net_value")):
            return r, found
    return None, {}


def _to_float(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        s = re.sub(r"[^0-9.\-]", "", s)
        return float(s) if s else 0.0


def _to_date_str(value):
    """Normalizes an SO Date cell (a datetime, or a 'YYYY/MM/DD' string,
    or a 'DD/MM/YYYY'-ish string) to an ISO 'YYYY-MM-DD' string."""
    import datetime as _dt
    if isinstance(value, _dt.datetime):
        return value.date().isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    s = str(value or "").strip()
    if not s:
        return ""
    m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    return s


def parse_so_details(file_path):
    """Returns a dict: {ok, message, rows} where rows is a list of dicts:
      {so_number, so_date, material_name, material_code, qty, sales_unit,
       net_value, unit_cost, status}
    unit_cost is net_value / qty (0 if qty is 0), left for the caller to
    treat as the GST-exclusive taxable value per unit — the same
    convention this app already uses for a manually entered purchase.
    """
    try:
        import openpyxl
    except ImportError:
        return dict(ok=False, message="openpyxl is not installed on the server.", rows=[])

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        return dict(ok=False, message=f"Could not open this as an Excel (.xlsx) file: {e}", rows=[])

    try:
        sheet_name = None
        for name in wb.sheetnames:
            if "so" in name.strip().lower() and "detail" in name.strip().lower():
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        header_row, cols = _find_header_row(ws)
        if not header_row:
            return dict(ok=False,
                        message=f"Couldn't find a header row with 'Material Code', 'Quantity', and "
                                f"'Net Value' columns on sheet '{sheet_name}'. This may not be an "
                                f"'SO Details' export, or the column headers have changed.",
                        rows=[])

        rows = []
        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            def cell(key):
                c = cols.get(key)
                return ws.cell(row=r, column=c).value if c else None

            material_code = cell("material_code")
            if material_code is None or str(material_code).strip() == "":
                continue  # blank/spacer row
            qty = _to_float(cell("quantity"))
            net_value = _to_float(cell("net_value"))
            rows.append(dict(
                so_number=str(cell("so_number") or "").strip(),
                so_date=_to_date_str(cell("so_date")),
                material_name=str(cell("material") or "").strip(),
                material_code=str(material_code).strip(),
                qty=qty,
                sales_unit=str(cell("sales_unit") or "").strip(),
                net_value=round(net_value, 2),
                unit_cost=round(net_value / qty, 4) if qty else 0.0,
                status=str(cell("status") or "").strip(),
            ))
    finally:
        wb.close()

    if not rows:
        return dict(ok=False, message=f"Sheet '{sheet_name}' had a matching header row but no data rows under it.",
                    rows=[])

    return dict(ok=True, message=f"Parsed {len(rows)} line(s) from sheet '{sheet_name}'.", rows=rows)
